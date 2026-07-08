from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from services.models import PipelineResult


HEADERS = [
    "\u5e8f\u53f7",
    "\u65e5\u671f",
    "\u6b4c\u624b",
    "\u6f14\u51fa\u573a\u6240",
    "\u6b4c\u5355\u51fa\u73b0\u6b21\u6570",
    "\u6b4c\u5355\u4ee3\u8868\u6b4c\u66f2",
    "\u7f6e\u4fe1\u5ea6",
]


def write_matches_xlsx(result: PipelineResult, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u5339\u914d\u7ed3\u679c"
    sheet.append(HEADERS)
    for match in result.matches:
        sheet.append(
            [
                match.index,
                match.date_display,
                match.artist_name,
                match.venue,
                match.playlist_song_count,
                "\uff1b".join(match.sample_songs),
                match.confidence,
            ]
        )

    header_fill = PatternFill("solid", fgColor="F8AFC8")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    widths = [8, 18, 24, 20, 16, 54, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column == 6)

    if result.matches:
        table = Table(displayName="ConcertMatches", ref=f"A1:G{len(result.matches) + 1}")
        style = TableStyleInfo(name="TableStyleMedium14", showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
    sheet.freeze_panes = "A2"
    workbook.save(output_path)
    return output_path
